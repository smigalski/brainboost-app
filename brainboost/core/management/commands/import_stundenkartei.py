from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime, time
from pathlib import Path
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from core.models import CustomUser, Lesson, ProgressEntry, StudentProfile, TutorProfile


SUBJECT_KEYWORDS = [
    ("mathe", ("mathe", "mathematik", "algebra", "geometr", "bruch", "gleichung", "funktion", "ableitung", "integral", "stochastik")),
    ("deutsch", ("deutsch", "grammatik", "rechtschreibung", "aufsatz", "textanalyse", "interpretation", "gedicht", "lektuere")),
    ("englisch", ("englisch", "english", "vokabel", "speaking", "listening", "reading comprehension")),
    ("franzoesisch", ("franzoesisch", "franzosisch", "french")),
    ("spanisch", ("spanisch", "spanish", "espanol")),
    ("chemie", ("chemie", "chemistry", "chemisch")),
    ("biologie", ("biologie", "bio", "biology", "genetik", "zelle", "oekologie")),
    ("physik", ("physik", "physics", "mechanik", "elektr", "optik", "waermelehre")),
    ("erdkunde", ("erdkunde", "geographie", "geografie", "geo", "klimazone", "atlas")),
    ("geschichte", ("geschichte", "history", "weimar", "nationalsozialismus", "mittelalter", "antik")),
    ("informatik", ("informatik", "programmierung", "python", "java", "algorithm", "datenbank", "sql", "html", "css")),
    ("politik", ("politik", "povi", "powi", "sozialkunde", "demokratie", "bundestag", "wahl", "eu", "wirtschaftspolitik")),
    ("musik", ("musik", "music", "noten", "harmonie", "rhythmus")),
    ("sonstiges_mint", ("naturwissenschaft", "mint", "technik", "wissenschaft")),
    ("sonstiges_gesellschaft", ("gesellschaft", "wirtschaft", "ethik", "philosophie", "religion", "werte", "gemeinschaftskunde")),
    ("sonstiges_sprache", ("sprache", "sprachlich", "literatur", "latein", "kommunikation")),
]

LOCATION_KEYWORDS = [
    (Lesson.Ort.ONLINE, ("online", "bbb", "bigbluebutton", "zoom", "teams", "meet")),
    (Lesson.Ort.BIB_WOB, ("wolfsburg", "wob", "bibliothek wob")),
    (Lesson.Ort.BIB, ("braunschweig", "bibliothek", "bib")),
    (Lesson.Ort.ZUHAUSE_BRAIN, ("brainboost", "bei mir", "bei tutor", "kiara")),
    (Lesson.Ort.ZUHAUSE_STUDENT, ("bei schueler", "bei schüler", "bei student", "zuhause", "hausbesuch")),
]

NAME_ALIASES = {
    "philip": "philipp",
    "linja + freundin": "linja",
}


class Command(BaseCommand):
    help = (
        "Importiert rueckwirkende Lernfortschrittseintraege aus der Excel-Datei "
        "'Stundenkartei' in Lessons und ProgressEntry."
    )

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", help="Pfad zur Excel-Datei (.xlsx)")
        parser.add_argument("--sheet", help="Optionaler Sheet-Name")
        parser.add_argument(
            "--default-time",
            default="14:00",
            help="Uhrzeit fuer neu erzeugte Termine im Format HH:MM",
        )
        parser.add_argument(
            "--duration",
            type=int,
            default=60,
            help="Dauer in Minuten fuer neu erzeugte Termine",
        )
        parser.add_argument(
            "--tutor-username",
            help="Optionaler TutorIn-Username als Fallback oder fuer mehrdeutige Zuordnungen",
        )
        parser.add_argument(
            "--fallback-subject",
            default="sonstiges_gesellschaft",
            choices=[choice[0] for choice in Lesson._meta.get_field("fach").choices],
            help="Fach, wenn aus dem Lerninhalt nichts erkannt wird",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Fuehrt den Import testweise aus und rollt alle DB-Aenderungen danach zurueck",
        )

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("openpyxl ist nicht installiert.") from exc

        xlsx_path = Path(options["xlsx_path"]).expanduser()
        if not xlsx_path.exists():
            raise CommandError(f"Datei nicht gefunden: {xlsx_path}")

        try:
            default_time = time.fromisoformat(options["default_time"])
        except ValueError as exc:
            raise CommandError("--default-time muss im Format HH:MM angegeben werden.") from exc

        tutor_override = None
        if options.get("tutor_username"):
            tutor_override = (
                TutorProfile.objects.select_related("user")
                .filter(user__username__iexact=options["tutor_username"])
                .first()
            )
            if not tutor_override:
                raise CommandError("TutorIn fuer --tutor-username nicht gefunden.")

        workbook = load_workbook(filename=xlsx_path, data_only=True)
        sheet = workbook[options["sheet"]] if options.get("sheet") else workbook.active
        header_map = self._header_map(sheet)

        stats = Counter()
        errors: list[str] = []
        student_lookup = self._student_lookup()

        with transaction.atomic():
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                if self._row_is_empty(row):
                    continue
                try:
                    lesson_date = self._parse_date(row[header_map["Datum"]], row_number)
                    location_value = self._clean_cell(row[header_map["Ort"]])
                    content_value = self._clean_cell(row[header_map["Lerninhalt"]])
                    student_value = self._clean_cell(row[header_map["Schüler/Student"]])

                    if not student_value:
                        raise CommandError("Schüler/Student fehlt.")

                    student = self._resolve_student(student_lookup, student_value)
                    tutor = self._resolve_tutor(student, tutor_override)
                    subject = self._infer_subject(content_value, options["fallback_subject"])
                    location = self._map_location(location_value)
                    lesson, lesson_created = self._resolve_lesson(
                        student=student,
                        tutor=tutor,
                        lesson_date=lesson_date,
                        location=location,
                        subject=subject,
                        default_time=default_time,
                        duration=options["duration"],
                    )
                    if lesson_created:
                        stats["lessons_created"] += 1
                    else:
                        stats["lessons_reused"] += 1

                    progress_created = self._create_progress_entry(
                        lesson=lesson,
                        comment=content_value,
                        lesson_date=lesson_date,
                        default_time=default_time,
                    )
                    if progress_created:
                        stats["progress_created"] += 1
                    else:
                        stats["progress_skipped"] += 1
                except Exception as exc:
                    stats["errors"] += 1
                    errors.append(f"Zeile {row_number}: {exc}")

            if options["dry_run"]:
                transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS("Import abgeschlossen."))
        self.stdout.write(
            f"Lessons neu: {stats['lessons_created']}, wiederverwendet: {stats['lessons_reused']}, "
            f"Lernfortschritte neu: {stats['progress_created']}, uebersprungen: {stats['progress_skipped']}, "
            f"Fehler: {stats['errors']}"
        )
        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("Dry-Run: keine Aenderungen wurden gespeichert."))
        if errors:
            self.stdout.write(self.style.WARNING("Fehlerdetails:"))
            for error in errors[:50]:
                self.stdout.write(f"- {error}")

    def _header_map(self, sheet):
        headers = [self._clean_cell(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        required = {"Datum", "Ort", "Lerninhalt", "Schüler/Student"}
        missing = required.difference(headers)
        if missing:
            raise CommandError(f"Fehlende Spalten in Excel: {', '.join(sorted(missing))}")
        return {header: index for index, header in enumerate(headers)}

    def _student_lookup(self):
        lookup = defaultdict(list)
        for student in StudentProfile.objects.select_related("user"):
            aliases = {
                student.user.username,
                student.user.first_name,
                student.user.last_name,
                student.user.get_full_name(),
                f"{student.user.last_name} {student.user.first_name}".strip(),
                f"{student.user.last_name}, {student.user.first_name}".strip(", "),
            }
            for alias in aliases:
                normalized = self._normalize(alias)
                if normalized:
                    lookup[normalized].append(student)
        return lookup

    def _resolve_student(self, lookup, raw_name: str) -> StudentProfile:
        normalized = self._normalize(raw_name)
        normalized = NAME_ALIASES.get(normalized, normalized)
        candidates = lookup.get(normalized, [])
        if not candidates:
            simplified = self._simplify_student_name(normalized)
            simplified = NAME_ALIASES.get(simplified, simplified)
            candidates = lookup.get(simplified, [])
        if len(candidates) == 1:
            return candidates[0]
        if len(candidates) > 1:
            raise CommandError(f"Mehrdeutige SchülerIn-Zuordnung fuer '{raw_name}'.")
        raise CommandError(f"Keine SchülerIn fuer '{raw_name}' gefunden.")

    def _resolve_tutor(
        self, student: StudentProfile, tutor_override: TutorProfile | None
    ) -> TutorProfile:
        assigned = list(student.assigned_tutors.select_related("user"))
        if tutor_override:
            if not assigned or any(t.pk == tutor_override.pk for t in assigned):
                return tutor_override
            raise CommandError(
                f"TutorIn '{tutor_override.user.username}' ist {student.user.get_full_name()} nicht zugewiesen."
            )
        if len(assigned) == 1:
            return assigned[0]
        if not assigned:
            raise CommandError(
                f"Keine TutorIn-Zuordnung fuer {student.user.get_full_name()} vorhanden. Bitte --tutor-username verwenden."
            )
        raise CommandError(
            f"Mehrere TutorInnen fuer {student.user.get_full_name()} gefunden. Bitte --tutor-username verwenden."
        )

    def _resolve_lesson(
        self,
        *,
        student: StudentProfile,
        tutor: TutorProfile,
        lesson_date: date,
        location: str,
        subject: str,
        default_time: time,
        duration: int,
    ):
        candidates = Lesson.objects.filter(student=student, date=lesson_date).order_by("time")
        lesson = None
        location_matches = candidates.filter(ort=location)
        if location_matches.count() == 1:
            lesson = location_matches.first()
        elif candidates.count() == 1:
            lesson = candidates.first()
        elif candidates.count() > 1:
            subject_matches = candidates.filter(fach=subject)
            if subject_matches.count() == 1:
                lesson = subject_matches.first()
            else:
                raise CommandError(
                    f"Mehrere Termine fuer {student.user.get_full_name()} am {lesson_date:%d.%m.%Y} gefunden."
                )

        created = False
        if lesson is None:
            lesson = Lesson.objects.create(
                student=student,
                tutor=tutor,
                date=lesson_date,
                time=default_time,
                duration_minutes=duration,
                ort=location,
                fach=subject,
                status=Lesson.Status.COMPLETED,
            )
            created = True
        else:
            updates = []
            if lesson.ort != location:
                lesson.ort = location
                updates.append("ort")
            if lesson.fach != subject:
                lesson.fach = subject
                updates.append("fach")
            if lesson.status != Lesson.Status.COMPLETED and lesson_date <= timezone.localdate():
                lesson.status = Lesson.Status.COMPLETED
                updates.append("status")
            if updates:
                lesson.save(update_fields=updates)
        return lesson, created

    def _create_progress_entry(
        self,
        *,
        lesson: Lesson,
        comment: str,
        lesson_date: date,
        default_time: time,
    ) -> bool:
        normalized_comment = (comment or "").strip()
        if ProgressEntry.objects.filter(lesson=lesson, comment=normalized_comment).exists():
            return False

        entry = ProgressEntry.objects.create(
            lesson=lesson,
            comment=normalized_comment,
            rating=None,
        )
        created_at = timezone.make_aware(datetime.combine(lesson_date, default_time))
        ProgressEntry.objects.filter(pk=entry.pk).update(created_at=created_at)
        return True

    def _infer_subject(self, content: str, fallback_subject: str) -> str:
        normalized = self._normalize(content)
        for subject, keywords in SUBJECT_KEYWORDS:
            if any(keyword in normalized for keyword in keywords):
                return subject
        return fallback_subject

    def _map_location(self, raw_location: str) -> str:
        normalized = self._normalize(raw_location)
        for value, keywords in LOCATION_KEYWORDS:
            if any(keyword in normalized for keyword in keywords):
                return value
        return Lesson.Ort.ZUHAUSE_STUDENT

    def _parse_date(self, value, row_number: int) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = self._clean_cell(value)
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        raise CommandError(f"Ungueltiges Datum '{text}' in Zeile {row_number}.")

    def _row_is_empty(self, row) -> bool:
        return all(self._clean_cell(value) == "" for value in row)

    def _clean_cell(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _normalize(self, value: str) -> str:
        cleaned = self._clean_cell(value).casefold()
        replacements = {
            "ä": "ae",
            "ö": "oe",
            "ü": "ue",
            "ß": "ss",
        }
        for source, target in replacements.items():
            cleaned = cleaned.replace(source, target)
        cleaned = re.sub(r"\s+", " ", cleaned)
        return cleaned

    def _simplify_student_name(self, value: str) -> str:
        cleaned = re.sub(r"\(.*?\)", "", value)
        cleaned = cleaned.split("+", 1)[0]
        return self._normalize(cleaned)
